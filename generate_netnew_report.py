#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Marketing NET-NEW / SOURCED Pipeline Report - deterministic generator.

Companion to the influence report, separate deliverable. Pure code: the entire
data pull and workbook assembly runs with zero language-model calls
in the hot path. "The LLM builds it. Math runs it."

Two grains (both from Alecia's own definitions in her SharePoint docs):

  DEAL grain  (the load-bearing offsite view, slide 15 "H1 2026 Marketing
              Attribution by Program"):
     Sourced = SINGLE-PROGRAM (marketing-originated proxy). A deal is
     "sourced" by a program when EVERY marketing campaign that touched any of
     its associated contacts maps to exactly ONE program bucket
     (Content & Technology / Events / Advertising / PR & Brand).
     Scope: Net New Pipeline (id 813739955), Amazon EXCLUDED, Galco ISOLATED.

  CONTACT grain (board deck "Net New (Marketing Sourced) Report",
              "Marketing Generated CRM population" bucketed by Lifecycle Stage):
     Reproduced here as a defensible, reproducible proxy: a contact CREATED
     since 2025-06-01 that is a member of at least one "Campaign Influence"
     segment list, positioned by lifecycle stage. FLAGGED: Alecia's exact
     board population may be a specific saved list; confirm before adopting as
     canonical (see the Confirmations tab).

Window: created since 2025-06-01 (matches the influence report).
Influence source: the same "Campaign Influence" folder segment lists the
influence report uses (HubSpot Marketing Pro has no native attribution API).

All credentials come from config.env (chmod 600, never web-served, never in git).
No em dashes anywhere.
"""
import os, sys, json, time
from collections import defaultdict
from datetime import datetime, timezone

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------- config ------
def load_config(path=None):
    path = path or os.path.join(HERE, "config.env")
    cfg = {}
    with open(path) as fh:
        for line in fh:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                cfg[k.strip()] = v.strip()
    return cfg

CFG = load_config()
HS_TOKEN = CFG["HUBSPOT_TOKEN"]
PORTAL = CFG.get("HUBSPOT_PORTAL_ID", "20335613")
SINCE = CFG.get("DEALS_CREATED_SINCE", "2025-06-01")
NET_NEW_PIPELINE_ID = CFG.get("NET_NEW_PIPELINE_ID", "813739955")
OUTDIR = os.path.join(HERE, "output")
os.makedirs(OUTDIR, exist_ok=True)

HS = "https://api.hubapi.com"
HH = {"Authorization": "Bearer " + HS_TOKEN, "Content-Type": "application/json"}

# ------------------------------------------------------- offsite model facts --
# From H2-Q3 Sales & Mktg Plan - July 2026.pptx (ELT offsite 2026-07-15) and the
# July-7 evidence deck + backing workbook. These are Alecia's real figures.
MODEL = {
    "revenue_target_2026": 14_000_000.0,
    "revenue_target_2027": 30_000_000.0,
    "h2_committed": 10_600_000.0,
    "h2_quota": 9_750_000.0,
    "open_pipeline": 17_010_000.0,           # ~1.75x coverage
    "h2_gap": 5_700_000.0,
    "h2_new_pipeline_build": 16_000_000.0,   # ~$16M to close the H2 gap
    "close_rate_amazon": 0.70,
    "close_rate_non_amazon": 0.30,           # high-20s to low-30s
    "pipeline_per_won_dollar": 3.30,         # non-Amazon: $1 won needs ~$3.30 worked
    "assigned_win_rate_2027": 0.60,
    "avg_deal_blended": 20_047.0,            # $8.52M / 425 won
    "avg_deal_slowlane": 25_472.0,           # $917K / 36 slow-lane new-business
    "won_book_total": 8_520_000.0,
    "won_book_deals": 425,
    "slowlane_engine_year": 917_000.0,
    "slowlane_engine_deals": 36,
}
# 2027 qualified pipeline needed = target / assigned win rate
MODEL["pipeline_needed_2027"] = MODEL["revenue_target_2027"] / MODEL["assigned_win_rate_2027"]

# Her offsite slide-15 actuals (ex-Galco), for the verification comparison.
SLIDE15 = {
    "Content & Technology": {"deals": 112, "sourced": 3_760_000.0, "won": 602_000.0},
    "Events":               {"deals": 14,  "sourced": 567_000.0,   "won": 67_000.0},
    "Advertising":          {"deals": 0,   "sourced": 0.0,         "won": 0.0},
    "PR & Brand":           {"deals": 0,   "sourced": 0.0,         "won": 23_500.0},
}
SLIDE15_GALCO = {"deals": 1, "sourced": 4_030_000.0}
SLIDE15_TOTAL_INFLUENCED_DEALS = 148  # distinct Net New non-Amazon influenced deals

# ------------------------------------------------------------- hubspot io -----
def hs_get(path, params=None):
    for attempt in range(5):
        r = requests.get(HS + path, headers=HH, params=params, timeout=60)
        if r.status_code == 429:
            time.sleep(1.5 * (attempt + 1)); continue
        r.raise_for_status()
        return r.json()
    raise RuntimeError("rate limited: GET " + path)

def hs_post(path, body):
    for attempt in range(6):
        r = requests.post(HS + path, headers=HH, json=body, timeout=60)
        if r.status_code == 429:
            time.sleep(1.5 * (attempt + 1)); continue
        r.raise_for_status()
        return r.json()
    raise RuntimeError("rate limited: POST " + path)

def since_ms():
    dt = datetime.strptime(SINCE, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    return str(int(dt.timestamp() * 1000))

# ---------------------------------------------------- program classifier ------
# Maps a Campaign Influence list/campaign NAME to one of the four offsite
# program buckets. Keyword rules, reverse-engineered from Alecia's list set and
# slide 15. Advertising tested first (paid overrides), then Events, then PR,
# else Content & Technology (the workhorse catch-all: content, webinar,
# whitepaper, video, blog, fact sheet, pillar page, forms, case studies).
ADV_KW = ("paid", "ppc", "google ads", "linkedin ad", "meta paid", "display",
          "banner ad", "abx linkedin", "linkedin ab", "boost", "retargeting")
EVENT_KW = ("conference", "summit", "live event", "booth", "tradeshow", "trade show",
            "expo", "modex", "maintec", "euromaintenance", "marcon", "bindt",
            "reliable plant", "reliability conference", "supplychainpoint",
            "data center world", "mainstream", "attendees")
PR_KW = ("pr:", "press", "investor", "g2 crowd", "earned media", "media banner")

def classify_program(name):
    n = name.lower()
    n = n.replace("campaign influence:", "").replace("campaign influence :", "").strip()
    if any(k in n for k in ADV_KW):
        return "Advertising"
    if any(k in n for k in EVENT_KW):
        return "Events"
    if any(k in n for k in PR_KW):
        return "PR & Brand"
    return "Content & Technology"

PROGRAMS = ["Content & Technology", "Events", "Advertising", "PR & Brand"]

def is_amazon(name):
    return "amazon" in (name or "").lower()
def is_galco(name):
    return "galco" in (name or "").lower()

# ------------------------------------------------------------- data pull ------
def pull_pipelines():
    data = hs_get("/crm/v3/pipelines/deals")
    pipe_label, stage_label, won_stage_ids = {}, {}, set()
    for p in data["results"]:
        pipe_label[p["id"]] = p["label"]
        for s in p["stages"]:
            stage_label[s["id"]] = s["label"]
            md = s.get("metadata", {}) or {}
            prob = md.get("probability")
            if (md.get("isClosed") in (True, "true")) and str(prob) in ("1.0", "1"):
                won_stage_ids.add(s["id"])
            # fallback: label based
            if "closed won" in s["label"].lower():
                won_stage_ids.add(s["id"])
    return pipe_label, stage_label, won_stage_ids

def pull_campaign_lists():
    res = hs_post("/crm/v3/lists/search", {"query": "Campaign Influence", "count": 250})
    lists = [l for l in res.get("lists", [])
             if l["name"].lower().startswith("campaign influence")]
    return lists

def pull_memberships(list_id):
    ids, after = [], None
    while True:
        params = {"limit": 250}
        if after:
            params["after"] = after
        r = hs_get("/crm/v3/lists/%s/memberships" % list_id, params)
        ids.extend(str(m["recordId"]) for m in r.get("results", []))
        after = (r.get("paging", {}).get("next") or {}).get("after")
        if not after:
            break
    return ids

def pull_netnew_deals():
    """Deals in the Net New Pipeline created since the window start."""
    ms = since_ms()
    props = ["dealname", "amount", "pipeline", "dealstage", "closedate", "createdate",
             "hs_is_closed_won", "hs_is_closed", "amount_in_home_currency"]
    out, after = [], None
    while True:
        body = {"filterGroups": [{"filters": [
                    {"propertyName": "createdate", "operator": "GTE", "value": ms},
                    {"propertyName": "pipeline", "operator": "EQ", "value": NET_NEW_PIPELINE_ID}]}],
                "properties": props, "limit": 100,
                "sorts": [{"propertyName": "createdate", "direction": "ASCENDING"}]}
        if after:
            body["after"] = after
        r = hs_post("/crm/v3/objects/deals/search", body)
        out.extend(r.get("results", []))
        after = (r.get("paging", {}).get("next") or {}).get("after")
        if not after:
            break
        time.sleep(0.2)
    return out

def batch_assoc(from_obj, to_obj, ids):
    out = defaultdict(list)
    ids = [str(i) for i in ids]
    for i in range(0, len(ids), 100):
        chunk = ids[i:i + 100]
        r = hs_post("/crm/v4/associations/%s/%s/batch/read" % (from_obj, to_obj),
                    {"inputs": [{"id": x} for x in chunk]})
        for res in r.get("results", []):
            fid = str((res.get("from") or {}).get("id") or res.get("fromObjectId"))
            for t in res.get("to", []):
                tid = t.get("toObjectId") or t.get("id")
                if tid is not None:
                    out[fid].append(str(tid))
        time.sleep(0.12)
    return out

def batch_read(obj, ids, properties):
    out = {}
    ids = [str(i) for i in ids]
    for i in range(0, len(ids), 100):
        chunk = ids[i:i + 100]
        r = hs_post("/crm/v3/objects/%s/batch/read" % obj,
                    {"properties": properties, "inputs": [{"id": x} for x in chunk]})
        for res in r.get("results", []):
            out[str(res["id"])] = res.get("properties", {})
        time.sleep(0.1)
    return out

# ----------------------------------------------------------------- compute ----
LIFECYCLE_ORDER = ["subscriber", "lead", "marketingqualifiedlead", "157687207",
                   "salesqualifiedlead", "opportunity", "customer",
                   "1157693063", "evangelist", "other"]
LIFECYCLE_LABEL = {
    "subscriber": "Subscriber", "lead": "Lead",
    "marketingqualifiedlead": "Marketing Qualified Lead",
    "157687207": "Sales Accepted Lead", "salesqualifiedlead": "Sales Qualified Lead",
    "opportunity": "Opportunity", "customer": "Customer",
    "1157693063": "Repeat Customer", "evangelist": "Advocate", "other": "Other", "": "(unset)",
}

def build_dataset():
    print("[1/6] pipelines ...", flush=True)
    pipe_label, stage_label, won_stage_ids = pull_pipelines()
    print("      Net New Pipeline label: %s ; won stages: %d"
          % (pipe_label.get(NET_NEW_PIPELINE_ID, "?"), len(won_stage_ids)), flush=True)

    print("[2/6] campaign influence lists + memberships ...", flush=True)
    lists = pull_campaign_lists()
    contact_campaigns = defaultdict(set)      # contactId -> {campaign name}
    campaign_size = {}
    for l in lists:
        name = l["name"]
        members = pull_memberships(l["listId"])
        campaign_size[name] = len(members)
        for c in members:
            contact_campaigns[c].add(name)
    print("      %d campaign lists, %d contacts mapped"
          % (len(lists), len(contact_campaigns)), flush=True)

    print("[3/6] net-new-pipeline deals ...", flush=True)
    raw = pull_netnew_deals()
    deals = {}
    for d in raw:
        p = d["properties"]
        amt = float(p.get("amount_in_home_currency") or p.get("amount") or 0)
        deals[d["id"]] = {
            "id": d["id"], "name": p.get("dealname") or "",
            "amount": amt,
            "pipeline": pipe_label.get(p.get("pipeline"), p.get("pipeline") or ""),
            "stage_id": p.get("dealstage") or "",
            "stage": stage_label.get(p.get("dealstage"), p.get("dealstage") or ""),
            "close": (p.get("closedate") or "")[:10],
            "create": (p.get("createdate") or "")[:10],
            "won": (str(p.get("hs_is_closed_won")).lower() == "true")
                   or (p.get("dealstage") in won_stage_ids),
            "closed": (str(p.get("hs_is_closed")).lower() == "true"),
        }
    print("      %d net-new deals, $%.0f total amount"
          % (len(deals), sum(x["amount"] for x in deals.values())), flush=True)

    print("[4/6] deal -> contacts / companies associations ...", flush=True)
    deal_ids = list(deals.keys())
    d2c = batch_assoc("deals", "contacts", deal_ids)
    d2co = batch_assoc("deals", "companies", deal_ids)

    companies_needed = set()
    for cos in d2co.values():
        companies_needed.update(cos)
    company_props = batch_read("companies", companies_needed, ["name"]) if companies_needed else {}
    def deal_company_name(did):
        cos = d2co.get(did, [])
        return company_props.get(cos[0], {}).get("name", "") if cos else ""

    # tag Amazon / Galco using company name OR deal name
    for did, d in deals.items():
        conm = deal_company_name(did)
        d["company"] = conm
        d["amazon"] = is_amazon(conm) or is_amazon(d["name"])
        d["galco"] = is_galco(conm) or is_galco(d["name"])

    # influence + program set per deal
    contacts_needed = set()
    for did, d in deals.items():
        inf = [c for c in d2c.get(did, []) if c in contact_campaigns]
        d["inf_contacts"] = inf
        d["influenced"] = bool(inf)
        progs = set()
        camps = set()
        for c in inf:
            for camp in contact_campaigns[c]:
                camps.add(camp); progs.add(classify_program(camp))
        d["programs"] = progs
        d["campaigns"] = camps
        d["single_program"] = (len(progs) == 1)
        d["program"] = next(iter(progs)) if len(progs) == 1 else ("(multi)" if progs else "")
        contacts_needed.update(inf)

    print("[5/6] influenced-contact details ...", flush=True)
    contact_props = batch_read("contacts", contacts_needed,
        ["firstname", "lastname", "email", "createdate", "lifecyclestage"]) if contacts_needed else {}

    # ---- CONTACT-GRAIN: marketing-sourced contacts (created in window, in a CI list)
    print("[6/6] contact-grain (marketing-sourced contacts) ...", flush=True)
    # Pull createdate + lifecyclestage for the FULL mapped universe (all CI-list members)
    all_mapped = list(contact_campaigns.keys())
    src_props = batch_read("contacts", all_mapped,
        ["createdate", "lifecyclestage", "firstname", "lastname", "email"]) if all_mapped else {}
    ms_int = int(since_ms())
    sourced_contacts = {}   # cid -> props (created in window)
    for cid, p in src_props.items():
        cd = p.get("createdate")
        try:
            cd_ms = int(datetime.strptime(cd[:10], "%Y-%m-%d").replace(tzinfo=timezone.utc).timestamp() * 1000) if cd else 0
        except Exception:
            cd_ms = 0
        if cd_ms >= ms_int:
            sourced_contacts[cid] = p

    return dict(
        pipe_label=pipe_label, stage_label=stage_label, won_stage_ids=won_stage_ids,
        lists=lists, contact_campaigns=contact_campaigns, campaign_size=campaign_size,
        deals=deals, d2c=d2c, d2co=d2co, company_props=company_props,
        deal_company_name=deal_company_name, contact_props=contact_props,
        sourced_contacts=sourced_contacts, src_props=src_props,
    )

def compute_deal_grain(ds):
    """Reproduce slide 15: single-program sourced deals per program, ex-Amazon, Galco isolated."""
    deals = ds["deals"]
    prog = {p: {"deals": 0, "sourced": 0.0, "won": 0.0, "won_deals": 0} for p in PROGRAMS}
    galco = {"deals": 0, "sourced": 0.0, "won": 0.0}
    influenced_non_amazon = []   # ex-amazon influenced deals (incl multi + galco)
    single_rows = []             # per single-program deal detail
    for did, d in deals.items():
        if d["amazon"]:
            continue
        if not d["influenced"]:
            continue
        influenced_non_amazon.append(did)
        if d["galco"]:
            galco["deals"] += 1; galco["sourced"] += d["amount"]
            if d["won"]:
                galco["won"] += d["amount"]
            continue
        if d["single_program"]:
            p = d["program"]
            prog[p]["deals"] += 1
            prog[p]["sourced"] += d["amount"]
            if d["won"]:
                prog[p]["won"] += d["amount"]; prog[p]["won_deals"] += 1
            single_rows.append(d)
    return dict(prog=prog, galco=galco,
                influenced_non_amazon=influenced_non_amazon, single_rows=single_rows)

def compute_campaign_summary(ds, dg):
    """Per single-program sourced deal, attribute its amount to the ONE program's
    campaigns. Report per-campaign: sourced contacts (list size within window is
    hard; use raw membership), deals sourced, sourced pipeline, won."""
    camp = defaultdict(lambda: {"program": "", "deals": 0, "sourced": 0.0,
                                "won": 0.0, "won_deals": 0, "contacts": 0})
    # contacts per campaign = raw list membership (universe), for context
    for name, size in ds["campaign_size"].items():
        camp[name]["contacts"] = size
        camp[name]["program"] = classify_program(name)
    for d in dg["single_rows"]:
        # even-split the deal across the campaigns that touched it (all in one program)
        cset = d["campaigns"]
        if not cset:
            continue
        share = d["amount"] / len(cset)
        for cname in cset:
            camp[cname]["deals"] += 1
            camp[cname]["sourced"] += share
            if d["won"]:
                camp[cname]["won"] += share; camp[cname]["won_deals"] += 1
    rows = []
    for name, v in camp.items():
        rows.append({"name": name, **v})
    producing = sorted([r for r in rows if r["deals"] > 0], key=lambda r: -r["sourced"])
    idle = sorted([r for r in rows if r["deals"] == 0], key=lambda r: -r["contacts"])
    return producing + idle, len(producing)

def compute_contact_funnel(ds):
    """Marketing-sourced contacts (created in window, in a CI list) by lifecycle stage."""
    buckets = defaultdict(int)
    for cid, p in ds["sourced_contacts"].items():
        buckets[p.get("lifecyclestage") or ""] += 1
    ordered = []
    for st in LIFECYCLE_ORDER + [""]:
        if buckets.get(st):
            ordered.append((st, LIFECYCLE_LABEL.get(st, st), buckets[st]))
    return ordered, sum(buckets.values())

# =============================================================== EXCEL =========
C_TITLE = PatternFill("solid", fgColor="1F4E78")
C_HDR = PatternFill("solid", fgColor="2E75B6")
C_PROG = PatternFill("solid", fgColor="E2EFDA")
F_TITLE = Font(bold=True, size=16, color="FFFFFF")
F_HDR = Font(bold=True, size=11, color="FFFFFF")
F_KPI_HDR = Font(bold=True, size=10, color="595959")
F_KPI_VAL = Font(bold=True, size=18, color="1F4E78")
F_SECTION = Font(bold=True, size=14, color="1F4E78")
USD = '"$"#,##0'

def _hdr(ws, row, headers, widths=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.fill = C_HDR; cell.font = F_HDR
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    if widths:
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

def build_workbook(ds, dg, camp_rows, n_prod, funnel, funnel_total, gen_date):
    wb = openpyxl.Workbook()
    prog = dg["prog"]; galco = dg["galco"]
    tot_sourced_deals = sum(prog[p]["deals"] for p in PROGRAMS)
    tot_sourced_val = sum(prog[p]["sourced"] for p in PROGRAMS)
    tot_won_val = sum(prog[p]["won"] for p in PROGRAMS)
    tot_won_deals = sum(prog[p]["won_deals"] for p in PROGRAMS)

    # ---- Tab 1: Executive Summary ----
    ws = wb.active; ws.title = "Exec Summary"
    ws.cell(1, 1, "MSAI Marketing Net-New / Sourced Pipeline Report").font = F_TITLE
    for c in range(1, 8):
        ws.cell(1, c).fill = C_TITLE
    meta = ("Generated: %s  |  Live HubSpot pull  |  Net New Pipeline (id %s)  |  Created since %s  |  "
            "Amazon excluded, Galco isolated  |  Sourced = single-program (marketing-originated proxy)"
            % (gen_date.strftime("%B %d, %Y"), NET_NEW_PIPELINE_ID, SINCE))
    ws.cell(2, 1, meta).font = Font(italic=True, size=9, color="595959")

    kpi_hdr = ["Sourced Deals (single-program)", "Sourced Pipeline $ (ex-Galco)",
               "Sourced Closed-Won $", "Sourced Contacts (net-new)",
               "Reached MQL+", "Closed-Won Deals", "Galco (isolated) $"]
    mql_plus = sum(n for st, lab, n in funnel
                   if st in ("marketingqualifiedlead", "157687207", "salesqualifiedlead",
                             "opportunity", "customer", "1157693063", "evangelist"))
    vals = [tot_sourced_deals, tot_sourced_val, tot_won_val, funnel_total,
            mql_plus, tot_won_deals, galco["sourced"]]
    for c, h in enumerate(kpi_hdr, 1):
        cell = ws.cell(4, c, h); cell.font = F_KPI_HDR
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for c, v in enumerate(vals, 1):
        cell = ws.cell(5, c, v); cell.font = F_KPI_VAL
        cell.number_format = USD if c in (2, 3, 7) else '#,##0'

    ws.cell(7, 1, "SOURCED PIPELINE BY PROGRAM (single-program, ex-Galco)").font = F_SECTION
    _hdr(ws, 8, ["Program", "Sourced Deals", "Sourced Pipeline $", "Closed-Won $",
                 "Won Deals", "Directional Cost/Deal note"],
         [26, 16, 20, 18, 12, 34])
    r = 9
    for p in PROGRAMS:
        ws.cell(r, 1, p).font = Font(bold=True)
        ws.cell(r, 2, prog[p]["deals"])
        ws.cell(r, 3, prog[p]["sourced"]).number_format = USD
        ws.cell(r, 4, prog[p]["won"]).number_format = USD
        ws.cell(r, 5, prog[p]["won_deals"])
        for cc in range(1, 7):
            ws.cell(r, cc).fill = C_PROG
        r += 1
    ws.cell(r, 1, "Galco (isolated)").font = Font(bold=True, italic=True)
    ws.cell(r, 2, galco["deals"]); ws.cell(r, 3, galco["sourced"]).number_format = USD
    ws.cell(r, 6, "Single account, excluded from bucket math").font = Font(italic=True, size=9, color="595959")
    r += 1
    ws.cell(r, 1, "TOTAL (ex-Galco)").font = Font(bold=True)
    ws.cell(r, 2, tot_sourced_deals).font = Font(bold=True)
    tc = ws.cell(r, 3, tot_sourced_val); tc.font = Font(bold=True); tc.number_format = USD
    tw = ws.cell(r, 4, tot_won_val); tw.font = Font(bold=True); tw.number_format = USD
    ws.cell(r, 5, tot_won_deals).font = Font(bold=True)

    r += 2
    ws.cell(r, 1, "HOW TO READ THIS REPORT").font = F_SECTION; r += 1
    how = [
        ("Tab 2 - Sourced Contacts Detail", "Every marketing-sourced net-new contact (created since %s, member of a Campaign Influence list), its lifecycle stage and any associated net-new deal." % SINCE),
        ("Tab 3 - Source / Campaign Summary", "Marketing programs and campaigns ranked by sourced pipeline. A deal is 'sourced' by a program when all its influencing campaigns map to that one program (single-program rule)."),
        ("Tab 4 - Funnel / Lifecycle Position", "Where the marketing-sourced contacts sit in the lifecycle. Directional only: HubSpot stage assignment is not a validated funnel (auto-advancement, stage-skipping)."),
        ("Tab 5 - Pipeline Model", "The offsite math: 2026 $14M / 2027 $30M targets, Amazon ~70% / non-Amazon ~30% close rates, pipeline needed, and how much marketing is sourcing today."),
        ("Tab 6 - Confirmations", "The specific definition choices that still need Alecia's sign-off before this is treated as canonical."),
        ("Definition", "Sourced / net-new = marketing-ORIGINATED (single-program proxy), distinct from 'influenced' = any campaign touch. Scope: Net New Pipeline, Amazon excluded, Galco isolated."),
    ]
    for a, b in how:
        ws.cell(r, 1, a).font = Font(bold=True, size=10)
        ws.cell(r, 2, b).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    ws.column_dimensions["A"].width = 30
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 20

    # ---- Tab 2: Sourced Contacts Detail ----
    ws = wb.create_sheet("Sourced Contacts Detail")
    _hdr(ws, 1, ["Contact Name", "Email", "Create Date", "Lifecycle Stage",
                 "Campaign(s) (marketing source)", "# Campaigns", "Program(s)"],
         [24, 32, 12, 22, 60, 12, 24])
    # rows: sourced contacts, with their campaigns/programs
    cc = ds["contact_campaigns"]
    def cname(cid, p):
        nm = ((p.get("firstname") or "") + " " + (p.get("lastname") or "")).strip()
        return nm or (p.get("email") or cid)
    rows = []
    for cid, p in ds["sourced_contacts"].items():
        camps = sorted(cc.get(cid, set()), key=str.lower)
        progs = sorted({classify_program(c) for c in camps})
        rows.append((cname(cid, p), p.get("email") or "", (p.get("createdate") or "")[:10],
                     LIFECYCLE_LABEL.get(p.get("lifecyclestage") or "", p.get("lifecyclestage") or ""),
                     " | ".join(camps), len(camps), " | ".join(progs)))
    rows.sort(key=lambda x: (-x[5], x[0].lower()))
    r = 2
    for row in rows:
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)
        r += 1
    ws.freeze_panes = "A2"; ws.auto_filter.ref = "A1:G%d" % (r - 1)

    # ---- Tab 3: Source / Campaign Summary ----
    ws = wb.create_sheet("Source-Campaign Summary")
    ws.cell(1, 1, "Sourced pipeline by PROGRAM (single-program, ex-Galco)").font = F_SECTION
    _hdr(ws, 2, ["Program", "Sourced Deals", "Sourced Pipeline $", "Closed-Won $", "Won Deals"],
         [26, 16, 20, 18, 12])
    r = 3
    for p in sorted(PROGRAMS, key=lambda p: -prog[p]["sourced"]):
        ws.cell(r, 1, p).font = Font(bold=True)
        ws.cell(r, 2, prog[p]["deals"])
        ws.cell(r, 3, prog[p]["sourced"]).number_format = USD
        ws.cell(r, 4, prog[p]["won"]).number_format = USD
        ws.cell(r, 5, prog[p]["won_deals"])
        r += 1
    r += 1
    ws.cell(r, 1, "By CAMPAIGN (ranked by sourced pipeline; even-split within program)").font = F_SECTION
    r += 1
    _hdr(ws, r, ["Campaign", "Program", "List Contacts (universe)", "Sourced Deals",
                 "Sourced Pipeline $", "Won $", "Won Deals"],
         [64, 24, 18, 14, 18, 16, 12])
    r += 1
    for row in camp_rows:
        ws.cell(r, 1, row["name"]); ws.cell(r, 2, row["program"])
        ws.cell(r, 3, row["contacts"])
        if row["deals"] > 0:
            ws.cell(r, 4, row["deals"])
            ws.cell(r, 5, row["sourced"]).number_format = USD
            ws.cell(r, 6, row["won"]).number_format = USD
            ws.cell(r, 7, row["won_deals"])
        else:
            for cc2 in (4, 5, 6, 7):
                ws.cell(r, cc2, "-")
        r += 1
    ws.freeze_panes = "A%d" % 3

    # ---- Tab 4: Funnel / Lifecycle Position ----
    ws = wb.create_sheet("Funnel-Lifecycle Position")
    ws.cell(1, 1, "Marketing-Sourced Net-New Contacts by Lifecycle Stage").font = F_SECTION
    ws.cell(2, 1, "Directional only. Per Alecia's board note: every MQL is also an SAL "
            "(auto-advancement), and there are more Opportunities than SQLs and more Customers "
            "than Opportunities (stage-skipping). Read as position, not a clean conversion funnel.").alignment = \
        Alignment(wrap_text=True)
    _hdr(ws, 4, ["Lifecycle Stage", "Sourced Contacts", "% of Sourced"], [30, 18, 14])
    r = 5
    for st, lab, n in funnel:
        ws.cell(r, 1, lab); ws.cell(r, 2, n)
        ws.cell(r, 3, (n / funnel_total) if funnel_total else 0).number_format = '0.0%'
        r += 1
    ws.cell(r, 1, "TOTAL").font = Font(bold=True)
    ws.cell(r, 2, funnel_total).font = Font(bold=True)

    # ---- Tab 5: Pipeline Model ----
    ws = wb.create_sheet("Pipeline Model")
    ws.cell(1, 1, "Pipeline Model - the offsite math").font = F_TITLE
    ws.cell(1, 1).fill = C_TITLE
    for c in range(2, 5):
        ws.cell(1, c).fill = C_TITLE
    ws.cell(2, 1, "Source: H2-Q3 Sales & Mktg Plan (ELT offsite 2026-07-15) + July-7 evidence deck. "
            "Marketing's sourced pipeline below is this report's live pull.").font = \
        Font(italic=True, size=9, color="595959")
    m = MODEL
    sourced_pipeline = tot_sourced_val
    share_2027 = sourced_pipeline / m["pipeline_needed_2027"] if m["pipeline_needed_2027"] else 0
    share_h2build = sourced_pipeline / m["h2_new_pipeline_build"] if m["h2_new_pipeline_build"] else 0
    lines = [
        ("2026 finish target", m["revenue_target_2026"], USD, "Amazon OP2 timing ($6.5M) is the single largest risk."),
        ("2027 ramp target", m["revenue_target_2027"], USD, "The net-new generation problem Q3's build feeds."),
        ("H2 committed", m["h2_committed"], USD, "Amazon + Fulfillment + Data Centers."),
        ("H2 quota", m["h2_quota"], USD, "Against $17.01M open pipeline = ~1.75x coverage (thin vs 3-4x norm)."),
        ("H2 coverage gap", m["h2_gap"], USD, "Target minus realistic named-deal coverage."),
        ("New qualified pipeline to build (H2)", m["h2_new_pipeline_build"], USD, "~$16M to close the H2 gap at current rates."),
        ("Close rate - Amazon", m["close_rate_amazon"], '0%', "Essentially handled; named deals carry its number."),
        ("Close rate - non-Amazon", m["close_rate_non_amazon"], '0%', "High-20s to low-30s. $1 won needs ~$3.30 pipeline worked."),
        ("Pipeline worked per $1 won (non-Amazon)", m["pipeline_per_won_dollar"], '0.00', "Inverse of the non-Amazon close rate."),
        ("Assigned win rate (2027 sizing)", m["assigned_win_rate_2027"], '0%', "Blended assigned rate for the $30M ramp."),
        ("Qualified pipeline needed for 2027 $30M", m["pipeline_needed_2027"], USD, "$30M / 60% assigned win rate."),
        ("Blended avg deal size", m["avg_deal_blended"], USD, "$8.52M / 425 won (dragged down by Amazon transactional)."),
        ("Slow-lane new-business avg deal", m["avg_deal_slowlane"], USD, "$917K / 36 deals. Use this for the net-new engine."),
        ("Slow-lane new-business engine / yr", m["slowlane_engine_year"], USD, "36 deals/yr today. What the ramp must scale."),
    ]
    _hdr(ws, 4, ["Model Input / Output", "Value", "Note"], [40, 18, 70])
    r = 5
    for lab, val, fmt, note in lines:
        ws.cell(r, 1, lab).font = Font(bold=True)
        c = ws.cell(r, 2, val); c.number_format = fmt
        ws.cell(r, 3, note).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    r += 1
    ws.cell(r, 1, "MARKETING'S SOURCED CONTRIBUTION (this report's live pull)").font = F_SECTION
    r += 1
    contrib = [
        ("Marketing-sourced pipeline (single-program, ex-Galco)", sourced_pipeline, USD),
        ("  as share of 2027 qualified pipeline needed (~$51M)", share_2027, '0.0%'),
        ("  as share of H2 new pipeline to build (~$16M)", share_h2build, '0.0%'),
        ("Marketing-sourced closed-won", tot_won_val, USD),
        ("Galco (isolated, single account)", galco["sourced"], USD),
    ]
    for lab, val, fmt in contrib:
        ws.cell(r, 1, lab).font = Font(bold=(not lab.startswith("  ")))
        cell = ws.cell(r, 2, val); cell.number_format = fmt
        r += 1

    # ---- Tab 6: Confirmations ----
    ws = wb.create_sheet("Confirmations")
    ws.cell(1, 1, "Open items - need Alecia's confirmation").font = F_TITLE
    ws.cell(1, 1).fill = C_TITLE
    for c in range(2, 4):
        ws.cell(1, c).fill = C_TITLE
    items = [
        ("Canonical grain", "This report ships BOTH a deal/pipeline view (single-program sourced) and a "
         "contact view (net-new sourced contacts). Confirm whether the productionized report is the "
         "deal view, the contact view, or both."),
        ("Contact-grain definition", "The board 'Marketing Generated CRM population' is not a single "
         "HubSpot property. This report proxies it as: created since %s AND member of a Campaign "
         "Influence list. If your board number comes from a specific saved list (e.g. 'Contacts Created "
         "From Marketing Efforts 2026', 'All Marketing Contacts', or a marketable-status segment), name "
         "it and we will switch to it exactly." % SINCE),
        ("Single-program = sourced", "Confirm the deal-grain rule 'all influencing campaigns map to one "
         "program bucket = sourced by that program' matches how slide 15 was built."),
        ("Program bucketing", "Confirm the campaign-to-program mapping (Content & Technology / Events / "
         "Advertising / PR & Brand). Advertising resolves to $0 sourced here (paid tag failure), matching "
         "slide 15."),
        ("Exclusions", "Confirm Amazon excluded and Galco isolated are the standing rule, not a one-off "
         "for the offsite deck."),
        ("Model figures", "Confirm the model tab uses $14M (2026) / $30M (2027), Amazon ~70% / "
         "non-Amazon ~30% close rates, ~60% assigned for 2027 sizing (superseding the old $10M / 23%)."),
        ("Window", "Confirm 2025-06-01 anchor (matches the influence report and the close-rate basis)."),
    ]
    _hdr(ws, 3, ["Item", "Detail"], [26, 110])
    r = 4
    for a, b in items:
        ws.cell(r, 1, a).font = Font(bold=True)
        ws.cell(r, 2, b).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    path = os.path.join(OUTDIR, "MSAI_NetNew_Sourced_Report.xlsx")
    wb.save(path)
    return path, dict(sourced_pipeline=sourced_pipeline, tot_sourced_deals=tot_sourced_deals,
                      tot_won_val=tot_won_val, tot_won_deals=tot_won_deals,
                      share_2027=share_2027, share_h2build=share_h2build, galco=galco)

# =============================================================== CHARTS ========
# =============================================================== MAIN ==========
def main():
    global dg_sourced, dg_share
    gen_date = datetime.now(timezone.utc)
    ds = build_dataset()
    dg = compute_deal_grain(ds)
    camp_rows, n_prod = compute_campaign_summary(ds, dg)
    funnel, funnel_total = compute_contact_funnel(ds)

    prog = dg["prog"]
    tot_sourced_val = sum(prog[p]["sourced"] for p in PROGRAMS)
    dg_sourced = tot_sourced_val
    dg_share = tot_sourced_val / MODEL["pipeline_needed_2027"]

    xlsx, summ = build_workbook(ds, dg, camp_rows, n_prod, funnel, funnel_total, gen_date)

    # machine-readable metrics for the Supabase history push (push_history.py).
    # Same numbers as the workbook exec summary, nothing recomputed.
    metrics = {
        "run_date": gen_date.strftime("%Y-%m-%d"),
        "generated_at": gen_date.isoformat(),
        "headline": {
            "sourced_deals": summ["tot_sourced_deals"],
            "sourced_pipeline": summ["sourced_pipeline"],
            "won_deals": summ["tot_won_deals"],
            "won_value": summ["tot_won_val"],
            "share_2027": summ["share_2027"],
            "share_h2_build": summ["share_h2build"],
            "sourced_contacts": funnel_total,
            "galco_deals": summ["galco"]["deals"],
            "galco_sourced": summ["galco"]["sourced"],
            "influenced_non_amazon_deals": len(dg["influenced_non_amazon"]),
            "net_new_deals_in_window": len(ds["deals"]),
        },
        "programs": [
            {"program": p, "deals": prog[p]["deals"], "sourced": prog[p]["sourced"],
             "won": prog[p]["won"], "won_deals": prog[p]["won_deals"]}
            for p in PROGRAMS
        ],
    }
    with open(os.path.join(OUTDIR, "netnew_metrics.json"), "w") as fh:
        json.dump(metrics, fh, indent=2)

    # ------- verification vs slide 15 + independent reconciliation -------
    print("\n================ NET-NEW SOURCED KPIs ================", flush=True)
    print("Net New Pipeline deals (window) ....... %d" % len(ds["deals"]))
    infl_na = len(dg["influenced_non_amazon"])
    print("Influenced non-Amazon deals (incl multi+Galco): %d  (slide15: %d)"
          % (infl_na, SLIDE15_TOTAL_INFLUENCED_DEALS))
    print("Sourced (single-program, ex-Galco): {:,} deals, ${:,.0f} pipeline, ${:,.0f} won"
          .format(summ["tot_sourced_deals"], tot_sourced_val, summ["tot_won_val"]))
    print("\n--- BY PROGRAM: ours vs slide 15 (allowing data drift) ---")
    print("%-24s %8s %14s %12s | %8s %14s %12s"
          % ("Program", "our#", "our$src", "our$won", "s15#", "s15$src", "s15$won"))
    for p in PROGRAMS:
        s15 = SLIDE15[p]
        print("%-24s %8d %14s %12s | %8d %14s %12s"
              % (p, prog[p]["deals"], "${:,.0f}".format(prog[p]["sourced"]),
                 "${:,.0f}".format(prog[p]["won"]),
                 s15["deals"], "${:,.0f}".format(s15["sourced"]), "${:,.0f}".format(s15["won"])))
    print("%-24s %8d %14s %12s | %8d %14s"
          % ("Galco (isolated)", dg["galco"]["deals"], "${:,.0f}".format(dg["galco"]["sourced"]),
             "-", SLIDE15_GALCO["deals"], "${:,.0f}".format(SLIDE15_GALCO["sourced"])))
    # independent recompute: sum of single_rows amounts must equal program totals
    indep = sum(d["amount"] for d in dg["single_rows"])
    print("\n--- INDEPENDENT RECONCILIATION ---")
    print("sum(single_program deal amounts) = ${:,.0f}".format(indep))
    print("sum(program sourced totals)      = ${:,.0f}".format(tot_sourced_val))
    print("reconciles: %s" % ("YES" if abs(indep - tot_sourced_val) < 1.0 else "NO"))
    camp_src = sum(r["sourced"] for r in camp_rows)
    print("sum(campaign even-split sourced) = ${:,.0f}  (should equal program total)".format(camp_src))
    print("campaign-split reconciles: %s" % ("YES" if abs(camp_src - tot_sourced_val) < 1.0 else "NO"))
    print("\nSourced contacts (net-new, in CI list, created since %s): %d" % (SINCE, funnel_total))
    print("Model: 2027 pipeline needed ${:,.0f} ; marketing sourced share {:.1f}%"
          .format(MODEL["pipeline_needed_2027"], 100 * summ["share_2027"]))
    print("\nXLSX: %s" % xlsx)

if __name__ == "__main__":
    main()

#!/usr/bin/env python3
r"""
MSAI Lead Pipeline SLA report.

Second report in the same package as generate_report.py. Same folder, same
.venv, same config.env, same HubSpot read-only token. It shares nothing with
the pipeline influence report at runtime and cannot affect it: different output
files, and it never reads or writes prior_snapshot.json.

WHAT IT PRODUCES  (both land in .\output)
    MSAI_Lead_Pipeline_SLA.xlsx   6-sheet workbook, the deliverable
    sla_snapshot.json             the same numbers as data, for the status page

READ-ONLY. It issues GET /crm/v3/owners, POST /crm/v3/objects/contacts/search
and POST /crm/v3/objects/contacts/batch/read. It never writes to HubSpot.

METHODOLOGY - this is the part that gets questioned, so it is stated plainly.
    Days in status is computed from the hs_lead_status PROPERTY HISTORY. For
    each contact we walk back from the newest history entry to the oldest entry
    in the current unbroken run of the current value. That timestamp is when the
    contact ENTERED its current status.

    It is deliberately NOT lastmodifieddate, which resets whenever anyone edits
    any part of the record and makes old leads look fresh. That is what produced
    the earlier, incorrect 26 percent figure for Qualified.

    It is deliberately NOT the custom field lead_status___last_updated_date,
    which is null for 209 of 293 Qualified contacts and does not reliably track
    entry into the current value.

    Days in lifecycle stage uses hs_v2_date_entered_current_stage, HubSpot's own
    computed "date entered current pipeline stage" property.

POPULATION
    Per-rep and per-stage timing covers contacts currently in lifecycle stage
    Sales Accepted Lead, MQL, SQL or Opportunity - the worked pipeline. The raw
    top-of-funnel Lead stage (about 29.5k largely unworked contacts) is excluded
    on purpose so per-rep averages reflect an actual working book rather than
    net-new volume.

    The SLA buckets themselves are portal-wide by lead status, because a
    Qualified or In Progress contact can also sit on Customer / Other /
    Evangelist stage.

SLA THRESHOLDS
    Read from config.env, not hardcoded. See SLA_DAYS_* there.
"""
import json
import os
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
BASE = "https://api.hubapi.com"


# ---------------------------------------------------------------- config ------
# Environment keys this module will accept as a stand-in for config.env.
# Scoped to known prefixes so os.environ at large never lands in CFG.
_ENV_PREFIXES = ("HUBSPOT_", "CAMPAIGN_", "DEALS_", "WINDSOR_", "NET_NEW_",
                 "SLA_DAYS_", "SUPABASE_", "MSAI_", "MKTG_")
_ENV_KEYS = ("PUBLISH_TARGET",)

def load_config(path=None):
    """Read config.env, then fill anything it does not set from the environment.

    The file is optional so this module can run on environment variables alone.
    When the file exists its values win, which keeps the CI path (the workflow
    writes config.env from secrets) behaving exactly as it did before."""
    path = path or os.path.join(HERE, "config.env")
    cfg = {}
    for k, v in os.environ.items():
        if k.startswith(_ENV_PREFIXES) or k in _ENV_KEYS:
            cfg[k] = v
    if os.path.exists(path):
        with open(path) as fh:
            for line in fh:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    cfg[k.strip()] = v.strip()
    return cfg


# Populated by init(), not at import time. See the note in generate_report.py.
CFG = {}
HS_TOKEN = None
PORTAL = "20335613"
OUTDIR = os.path.join(HERE, "output")
_READY = False


def _sla_days(key, default):
    """Thresholds live in config.env so the business can change the rule without
    touching code. A blank or unparseable value falls back to the default and
    says so, rather than silently reporting against a number nobody agreed to."""
    raw = (CFG.get(key, "") or "").strip()
    if raw == "":
        return float(default), False
    try:
        return float(raw), True
    except ValueError:
        print("WARNING: %s in config.env is not a number (%r). Using the default of %s days."
              % (key, raw, default), file=sys.stderr)
        return float(default), False


# Defaults stand until init() reads the configured thresholds.
SLA_DAYS = {
    "In Progress": 14.0,
    "Qualified": 14.0,
    "Awaiting Sales Qualification": 1.0,
}
SLA_FROM_CONFIG = {
    "In Progress": False,
    "Qualified": False,
    "Awaiting Sales Qualification": False,
}


def init(path=None, make_outdir=True):
    """Load config and populate module settings, including the SLA thresholds.
    Idempotent."""
    global CFG, HS_TOKEN, PORTAL, SLA_DAYS, SLA_FROM_CONFIG, _READY
    if _READY:
        return CFG
    CFG = load_config(path)
    HS_TOKEN = CFG.get("HUBSPOT_TOKEN") or ""
    if not HS_TOKEN:
        raise RuntimeError("HUBSPOT_TOKEN is not set. Put it in config.env next "
                           "to this script, or export it in the environment.")
    PORTAL = CFG.get("HUBSPOT_PORTAL_ID", "20335613")
    _ip, _ip_set = _sla_days("SLA_DAYS_IN_PROGRESS", 14)
    _q, _q_set = _sla_days("SLA_DAYS_QUALIFIED", 14)
    _a, _a_set = _sla_days("SLA_DAYS_AWAITING_SALES_QUALIFICATION", 1)
    SLA_DAYS = {
        "In Progress": _ip,
        "Qualified": _q,
        "Awaiting Sales Qualification": _a,
    }
    SLA_FROM_CONFIG = {
        "In Progress": _ip_set,
        "Qualified": _q_set,
        "Awaiting Sales Qualification": _a_set,
    }
    if make_outdir:
        os.makedirs(OUTDIR, exist_ok=True)
    _READY = True
    return CFG

STAGE_LABEL = {
    "subscriber": "Subscriber",
    "lead": "Lead",
    "marketingqualifiedlead": "MQL",
    "157687207": "Sales Accepted Lead",
    "salesqualifiedlead": "SQL",
    "opportunity": "Opportunity",
    "customer": "Customer",
    "1157693063": "Repeat Customer",
    "evangelist": "Advocate",
    "other": "Other",
}

PIPELINE_STAGES = ["157687207", "marketingqualifiedlead", "salesqualifiedlead", "opportunity"]

CONTACT_PROPS = [
    "hs_lead_status", "lifecyclestage", "hubspot_owner_id",
    "hs_v2_date_entered_current_stage", "hs_v2_time_in_current_stage",
    "createdate", "lastmodifieddate", "email", "firstname", "lastname", "company",
]


def threshold_label(days):
    if days == 1:
        return "24 hours"
    if float(days).is_integer():
        return "%d days" % int(days)
    return "%s days" % days


# ------------------------------------------------------------- hubspot -------
def _req(method, path, payload=None, params=None, retries=6):
    url = BASE + path
    if params:
        url += "?" + urllib.parse.urlencode(params)
    data = json.dumps(payload).encode() if payload is not None else None
    r = urllib.request.Request(
        url, data=data,
        headers={"Authorization": "Bearer %s" % HS_TOKEN, "Content-Type": "application/json"},
        method=method,
    )
    for _ in range(retries):
        try:
            return json.load(urllib.request.urlopen(r, timeout=30))
        except urllib.error.HTTPError as e:
            if e.code == 429:
                time.sleep(2)
                continue
            if e.code == 401:
                raise SystemExit(
                    "HubSpot returned 401 Unauthorized. The token in config.env is wrong, "
                    "expired or revoked. Nothing was written."
                )
            raise
    raise RuntimeError("failed after retries: " + path)


def fetch_owners():
    owners = {}
    after = None
    while True:
        params = {"limit": 100}
        if after:
            params["after"] = after
        r = _req("GET", "/crm/v3/owners", params=params)
        for o in r.get("results", []):
            name = ("%s %s" % (o.get("firstName") or "", o.get("lastName") or "")).strip()
            owners[str(o["id"])] = name or o.get("email") or str(o["id"])
        nxt = (r.get("paging") or {}).get("next")
        if not nxt:
            break
        after = nxt["after"]
    return owners


def search_contacts(prop, value):
    out = []
    after = None
    while True:
        payload = {
            "filterGroups": [{"filters": [{"propertyName": prop, "operator": "EQ", "value": value}]}],
            "properties": CONTACT_PROPS,
            "limit": 100,
        }
        if after:
            payload["after"] = after
        r = _req("POST", "/crm/v3/objects/contacts/search", payload)
        out.extend(r.get("results", []))
        nxt = (r.get("paging") or {}).get("next")
        if not nxt:
            break
        after = nxt["after"]
    return out


def fetch_status_history(ids):
    """Batch read, 50 ids per call. The earlier throwaway script did one GET per
    contact - about 412 sequential round trips. This is the same data in ~9."""
    out = {}
    CHUNK = 50
    for i in range(0, len(ids), CHUNK):
        chunk = ids[i:i + CHUNK]
        payload = {"inputs": [{"id": cid} for cid in chunk],
                   "propertiesWithHistory": ["hs_lead_status"]}
        r = _req("POST", "/crm/v3/objects/contacts/batch/read", payload)
        for res in r.get("results", []):
            out[res["id"]] = res.get("propertiesWithHistory", {}).get("hs_lead_status", [])
    return out


def entered_status(history, current_value):
    """Return (timestamp, source) for the moment the contact ENTERED its current
    lead status: the oldest entry in the current unbroken run of that value."""
    h = history or []
    if not h or h[0].get("value") != current_value:
        return None, None
    idx = 0
    while idx + 1 < len(h) and h[idx + 1].get("value") == current_value:
        idx += 1
    return h[idx].get("timestamp"), h[idx].get("sourceType")


def parse_ts(ts):
    if not ts:
        return None
    try:
        return datetime.fromisoformat(ts.replace("Z", "+00:00"))
    except ValueError:
        return None


def age_days(ts, now):
    dt = parse_ts(ts)
    if dt is None:
        return None
    return (now - dt).total_seconds() / 86400.0


def avg(lst):
    return round(sum(lst) / len(lst), 1) if lst else None


def median(lst):
    if not lst:
        return None
    s = sorted(lst)
    n = len(s)
    mid = n // 2
    return round(s[mid] if n % 2 else (s[mid - 1] + s[mid]) / 2.0, 1)


# ------------------------------------------------------------- snapshot ------
def build_snapshot():
    now = datetime.now(timezone.utc)
    print("[1/4] owners ...", flush=True)
    owners = fetch_owners()
    print("      %d owners" % len(owners), flush=True)

    def owner_of(props):
        oid = props.get("hubspot_owner_id")
        return owners.get(str(oid), "Unassigned") if oid else "Unassigned"

    print("[2/4] pipeline population (Sales Accepted Lead, MQL, SQL, Opportunity) ...", flush=True)
    stage_pop = {}
    for s in PIPELINE_STAGES:
        stage_pop[s] = search_contacts("lifecyclestage", s)
        print("      %-22s %d contacts" % (STAGE_LABEL.get(s, s), len(stage_pop[s])), flush=True)

    all_ids = []
    for contacts in stage_pop.values():
        all_ids.extend(c["id"] for c in contacts)
    all_ids = list(dict.fromkeys(all_ids))

    print("[3/4] lead status history for %d contacts (batches of 50) ..." % len(all_ids), flush=True)
    history = fetch_status_history(all_ids)

    records = []
    for stage, contacts in stage_pop.items():
        for c in contacts:
            p = c["properties"]
            status = p.get("hs_lead_status")
            stage_ts = p.get("hs_v2_date_entered_current_stage")
            status_ts, _src = entered_status(history.get(c["id"]), status) if status else (None, None)
            records.append({
                "id": c["id"],
                "stage_label": STAGE_LABEL.get(stage, stage),
                "status": status or "(blank)",
                "owner": owner_of(p),
                "stage_age_days": age_days(stage_ts, now),
                "status_age_days": age_days(status_ts, now),
            })

    print("[4/4] SLA buckets by lead status (portal-wide) ...", flush=True)
    headline = {}
    by_rep_sla = defaultdict(dict)
    stuck_buckets = []

    # One detail row per contact in the worked-pipeline population, seeded
    # before the SLA loop so the detail covers the whole book and not just the
    # three tracked lead-status buckets. Identity, owner and lifecycle-stage
    # timing populate for everyone; the SLA fields stay blank for contacts
    # whose lead_status is not one we track, which is correct rather than a gap.
    detail_by_id = {}
    for stage, contacts in stage_pop.items():
        for c in contacts:
            p = c["properties"]
            stage_raw = p.get("lifecyclestage") or stage
            detail_by_id[c["id"]] = {
                "contact_id": c["id"],
                "owner_id": p.get("hubspot_owner_id") or "",
                "status": "",
                "sla": None,
                "name": (("%s %s" % (p.get("firstname") or "",
                                     p.get("lastname") or "")).strip()),
                "email": p.get("email") or "",
                "company": p.get("company") or "",
                "rep": owner_of(p),
                "stage": STAGE_LABEL.get(stage_raw, stage_raw),
                "entered": "",
                "src": "",
                "days_status": "",
                "over": "",
                "days_stage": (lambda d: round(d, 1) if d is not None else "")(
                    age_days(p.get("hs_v2_date_entered_current_stage"), now)),
            }

    for status, sla in SLA_DAYS.items():
        contacts = search_contacts("hs_lead_status", status)
        ids = [c["id"] for c in contacts]
        need = [i for i in ids if i not in history]
        if need:
            history.update(fetch_status_history(need))

        ages = []
        over = 0
        rep_over = defaultdict(int)
        rep_total = defaultdict(int)

        for c in contacts:
            p = c["properties"]
            rep = owner_of(p)
            ts, src = entered_status(history.get(c["id"]), status)
            a = age_days(ts, now)
            rep_total[rep] += 1
            is_over = a is not None and a > sla
            if a is not None:
                ages.append(a)
                if is_over:
                    over += 1
                    rep_over[rep] += 1
            stage_raw = p.get("lifecyclestage") or ""
            # The lead-status search is portal-wide, so a handful of these sit
            # outside the four pipeline stages and have no seeded row yet.
            row = detail_by_id.setdefault(c["id"], {
                "contact_id": c["id"],
                "owner_id": p.get("hubspot_owner_id") or "",
                "name": (("%s %s" % (p.get("firstname") or "",
                                     p.get("lastname") or "")).strip()),
                "email": p.get("email") or "",
                "company": p.get("company") or "",
                "rep": rep,
                "stage": STAGE_LABEL.get(stage_raw, stage_raw),
                "days_stage": (lambda d: round(d, 1) if d is not None else "")(
                    age_days(p.get("hs_v2_date_entered_current_stage"), now)),
            })
            row.update({
                "status": status,
                "sla": sla,
                "entered": ts[:10] if ts else "(no history)",
                "src": src or "",
                "days_status": round(a, 1) if a is not None else "",
                "over": "YES" if is_over else "no",
            })

        total = len(contacts)
        headline[status] = {
            "total": total,
            "over": over,
            "pct_over": round(over / total * 100, 1) if total else 0,
            "avg_age_days": avg(ages),
            "median_age_days": median(ages),
            "sla_days": sla,
            "sla_label": threshold_label(sla),
            "sla_from_config": SLA_FROM_CONFIG[status],
        }
        print("      %-30s %d/%d over SLA (%.1f%%)"
              % (status, over, total, headline[status]["pct_over"]), flush=True)

        for rep, tot in rep_total.items():
            by_rep_sla[rep][status] = {"total": tot, "over": rep_over.get(rep, 0)}
            if rep_over.get(rep, 0) > 0:
                stuck_buckets.append({
                    "owner": rep, "status": status, "over": rep_over[rep],
                    "total": tot, "pct": round(rep_over[rep] / tot * 100, 1),
                })

    stuck_buckets.sort(key=lambda x: -x["over"])

    rep_agg = defaultdict(lambda: {"n": 0, "stage_ages": [], "status_ages": [], "last_change_ages": []})
    stage_agg = defaultdict(lambda: {"n": 0, "ages": []})
    overall_stage, overall_status, overall_last = [], [], []
    matrix = defaultdict(int)

    for r in records:
        rep_agg[r["owner"]]["n"] += 1
        if r["stage_age_days"] is not None:
            rep_agg[r["owner"]]["stage_ages"].append(r["stage_age_days"])
            stage_agg[r["stage_label"]]["ages"].append(r["stage_age_days"])
            overall_stage.append(r["stage_age_days"])
        if r["status_age_days"] is not None:
            rep_agg[r["owner"]]["status_ages"].append(r["status_age_days"])
            overall_status.append(r["status_age_days"])
        cands = [x for x in (r["stage_age_days"], r["status_age_days"]) if x is not None]
        if cands:
            last = min(cands)  # smaller age = more recent change
            rep_agg[r["owner"]]["last_change_ages"].append(last)
            overall_last.append(last)
        matrix[(r["stage_label"], r["status"], r["owner"])] += 1
        stage_agg[r["stage_label"]]["n"] += 1

    by_rep = {}
    for rep, agg in rep_agg.items():
        by_rep[rep] = {
            "pipeline_contacts": agg["n"],
            "avg_days_in_stage": avg(agg["stage_ages"]),
            "avg_days_in_status": avg(agg["status_ages"]),
            "avg_days_since_last_change": avg(agg["last_change_ages"]),
            "sla": dict(by_rep_sla.get(rep, {})),
        }
    for rep, sla in by_rep_sla.items():
        if rep not in by_rep:
            by_rep[rep] = {
                "pipeline_contacts": 0, "avg_days_in_stage": None,
                "avg_days_in_status": None, "avg_days_since_last_change": None,
                "sla": dict(sla),
            }

    by_stage = {k: {"total": v["n"], "avg_days_in_stage": avg(v["ages"])}
                for k, v in stage_agg.items()}

    snapshot = {
        "generated_at": now.isoformat(),
        "portal_id": PORTAL,
        "scope_note": (
            "Population = contacts currently in lifecycle stage Sales Accepted Lead, "
            "MQL, SQL, or Opportunity (the worked pipeline), plus the Qualified / In "
            "Progress / Awaiting Sales Qualification lead-status buckets portal-wide "
            "(a handful of those sit on Customer/Other/Evangelist stage). Raw "
            "top-of-funnel Lead stage (about 29.5k largely unworked contacts) is "
            "intentionally excluded from per-rep timing so the averages reflect "
            "actual working book, not net-new volume."
        ),
        "sla_thresholds": {k: v for k, v in SLA_DAYS.items()},
        "sla_thresholds_from_config": SLA_FROM_CONFIG,
        "headline": headline,
        "by_rep": by_rep,
        "by_stage": by_stage,
        "stuck_buckets": stuck_buckets,
        "matrix": [{"stage": k[0], "status": k[1], "owner": k[2], "count": v}
                   for k, v in matrix.items()],
        "overall": {
            "avg_days_in_stage": avg(overall_stage),
            "avg_days_in_status": avg(overall_status),
            "avg_days_since_last_change": avg(overall_last),
        },
        "methodology": {
            "lead_status_timestamp_property": (
                "hs_lead_status property history (propertiesWithHistory), oldest entry in "
                "the current unbroken run of the value - i.e. when the contact ENTERED its "
                "current status"),
            "lifecycle_stage_timestamp_property": "hs_v2_date_entered_current_stage",
            "rejected_property": (
                "lead_status___last_updated_date (custom field) - looks like a status-change "
                "stamp but is null for 209/293 Qualified contacts and does not reliably "
                "reflect entry into the CURRENT value"),
            "rejected_property_2": (
                "lastmodifieddate - resets on any edit to the record, which makes stale "
                "leads look fresh; this produced the earlier incorrect 26% for Qualified"),
        },
    }
    detail_rows = list(detail_by_id.values())
    return snapshot, detail_rows


# ------------------------------------------------------------- workbook ------
HDR = Font(bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="1F3864")
RED = PatternFill("solid", fgColor="F4CCCC")
BOLD = Font(bold=True)
TITLE = Font(bold=True, size=14)
_thin = Side(style="thin", color="D9D9D9")
B = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _header(ws, head, widths):
    for j, h in enumerate(head, 1):
        c = ws.cell(1, j, h)
        c.font = HDR
        c.fill = HFILL
        c.border = B
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def build_workbook(snap, rows, out_path):
    gen = parse_ts(snap["generated_at"]).strftime("%Y-%m-%d %H:%M UTC")
    wb = Workbook()

    # --- 1. How to read this -------------------------------------------------
    ws = wb.active
    ws.title = "How to read this"
    thr = "   ".join("%s: %s." % (k, threshold_label(v)) for k, v in SLA_DAYS.items())
    lines = [
        ("Lead Pipeline SLA - supporting data", TITLE),
        ("Source: HubSpot portal %s. Live pull %s." % (snap["portal_id"], gen), None),
        ("", None),
        ("What this measures", BOLD),
        ("How long each lead has been sitting in its current lead status, and whether that is past the SLA.", None),
        ("", None),
        ("How days-in-status is calculated (this is the part that gets questioned)", BOLD),
        ("For every contact we read the full change history of the Lead Status field and take the date the contact ENTERED its current status.", None),
        ("Days in status = today minus that entered date. If that is longer than the SLA, it is flagged Over SLA.", None),
        ("We do NOT use Last Modified Date. That field resets whenever anyone edits any part of the record, which makes old leads look fresh. That is what produced the earlier, incorrect 26 percent for Qualified.", None),
        ("We also do NOT use the custom field lead_status___last_updated_date. It is empty for 209 of 293 Qualified contacts.", None),
        ("", None),
        ("SLA thresholds in force for this run", BOLD),
        (thr, None),
        ("These are set in config.env on the machine that runs the report (SLA_DAYS_IN_PROGRESS, SLA_DAYS_QUALIFIED, SLA_DAYS_AWAITING_SALES_QUALIFICATION). Change the number there and re-run; no code change is needed.", None),
        ("", None),
        ("Who is counted", BOLD),
        (snap["scope_note"], None),
        ("", None),
        ("How to verify any number", BOLD),
        ("Open the All Contacts tab, filter by Status and by Over SLA = YES, and count the rows. It will match the Summary tab exactly.", None),
        ("Each row shows the entered date and the source of that status change (for example CRM_UI or a workflow), so every figure traces back to a real record.", None),
        ("", None),
        ("This report is read-only. It never writes anything back to HubSpot.", BOLD),
    ]
    for i, (t, f) in enumerate(lines, 1):
        c = ws.cell(i, 1, t)
        c.alignment = Alignment(wrap_text=False, vertical="top")
        if f:
            c.font = f
    ws.column_dimensions["A"].width = 120

    # --- 2. Summary by Status ------------------------------------------------
    ws = wb.create_sheet("Summary by Status")
    _header(ws, ["Lead Status", "SLA", "Total contacts", "Over SLA", "% Over SLA",
                 "Avg days in status", "Median days in status"],
            [30, 12, 15, 12, 13, 20, 22])
    r = 2
    for st, sla in SLA_DAYS.items():
        h = snap["headline"][st]
        vals = [st, threshold_label(sla), h["total"], h["over"],
                ("%.1f%%" % h["pct_over"]),
                h["avg_age_days"] if h["avg_age_days"] is not None else "",
                h["median_age_days"] if h["median_age_days"] is not None else ""]
        for j, v in enumerate(vals, 1):
            c = ws.cell(r, j, v)
            c.border = B
            if j == 5:
                c.font = BOLD
        r += 1

    # --- 3. Summary by Rep ---------------------------------------------------
    ws = wb.create_sheet("Summary by Rep")
    _header(ws, ["Rep / Owner", "In Progress over/total", "Qualified over/total",
                 "Awaiting SQL over/total", "Contacts in worked pipeline",
                 "Avg days in lifecycle stage", "Avg days in lead status",
                 "Avg days since last change"],
            [26, 20, 20, 22, 24, 24, 22, 24])

    def ot(sla_map, st):
        d = sla_map.get(st)
        if not d:
            return "-"
        return "%d/%d" % (d["over"], d["total"])

    r = 2
    for rep in sorted(snap["by_rep"]):
        d = snap["by_rep"][rep]
        vals = [rep,
                ot(d["sla"], "In Progress"),
                ot(d["sla"], "Qualified"),
                ot(d["sla"], "Awaiting Sales Qualification"),
                d["pipeline_contacts"],
                d["avg_days_in_stage"] if d["avg_days_in_stage"] is not None else "",
                d["avg_days_in_status"] if d["avg_days_in_status"] is not None else "",
                d["avg_days_since_last_change"] if d["avg_days_since_last_change"] is not None else ""]
        for j, v in enumerate(vals, 1):
            ws.cell(r, j, v).border = B
        r += 1
    ws.freeze_panes = "A2"

    # --- 4. Aging by Lifecycle Stage ----------------------------------------
    ws = wb.create_sheet("Aging by Lifecycle Stage")
    _header(ws, ["Lifecycle stage", "Contacts", "Avg days in stage"], [28, 14, 22])
    r = 2
    for st in sorted(snap["by_stage"], key=lambda k: -snap["by_stage"][k]["total"]):
        d = snap["by_stage"][st]
        for j, v in enumerate([st, d["total"],
                               d["avg_days_in_stage"] if d["avg_days_in_stage"] is not None else ""], 1):
            ws.cell(r, j, v).border = B
        r += 1
    r += 1
    ws.cell(r, 1, "Overall, worked pipeline").font = BOLD
    r += 1
    for label, key in [("Avg days in lifecycle stage", "avg_days_in_stage"),
                       ("Avg days in lead status", "avg_days_in_status"),
                       ("Avg days since last change", "avg_days_since_last_change")]:
        ws.cell(r, 1, label)
        ws.cell(r, 2, snap["overall"][key] if snap["overall"][key] is not None else "")
        r += 1

    # --- 5. Stage x Status x Rep --------------------------------------------
    ws = wb.create_sheet("Stage x Status x Rep")
    _header(ws, ["Lifecycle stage", "Lead status", "Rep / Owner", "Contacts"],
            [24, 30, 26, 12])
    r = 2
    for m in sorted(snap["matrix"], key=lambda x: (x["stage"], x["status"], -x["count"])):
        for j, v in enumerate([m["stage"], m["status"], m["owner"], m["count"]], 1):
            ws.cell(r, j, v).border = B
        r += 1
    ws.freeze_panes = "A2"

    # --- 6. All Contacts (detail) -------------------------------------------
    ws = wb.create_sheet("All Contacts (detail)")
    _header(ws, ["Lead Status", "SLA (days)", "Contact", "Email", "Company",
                 "Rep / Owner", "Lifecycle stage", "Entered current status",
                 "Status change source", "Days in status", "Over SLA?",
                 "Days in lifecycle stage"],
            [26, 11, 24, 30, 26, 22, 20, 20, 20, 14, 11, 20])
    det = sorted(rows, key=lambda x: (x["status"],
                                      -(x["days_status"] if isinstance(x["days_status"], (int, float)) else 0)))
    r = 2
    for x in det:
        vals = [x["status"], x["sla"], x["name"], x["email"], x["company"], x["rep"],
                x["stage"], x["entered"], x["src"], x["days_status"], x["over"], x["days_stage"]]
        for j, v in enumerate(vals, 1):
            c = ws.cell(r, j, v)
            c.border = B
            if j == 11 and v == "YES":
                c.fill = RED
                c.font = BOLD
        r += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:L%d" % (r - 1)

    wb.save(out_path)
    return out_path, len(det)


# ----------------------------------------------------------------- main ------
def main():
    init()
    t0 = time.time()
    snap, rows = build_snapshot()

    json_path = os.path.join(OUTDIR, "sla_snapshot.json")
    with open(json_path, "w") as fh:
        json.dump(snap, fh, indent=2)

    xlsx_path = os.path.join(OUTDIR, "MSAI_Lead_Pipeline_SLA.xlsx")
    xlsx_path, n_rows = build_workbook(snap, rows, xlsx_path)

    print("\n================ LEAD PIPELINE SLA ================")
    for st, h in snap["headline"].items():
        src = "config.env" if h["sla_from_config"] else "built-in default"
        print("%-30s %4d/%-4d over SLA (%5.1f%%)  avg %s d   SLA %s [%s]"
              % (st, h["over"], h["total"], h["pct_over"], h["avg_age_days"],
                 h["sla_label"], src))
    print("Reps ........................ %d" % len(snap["by_rep"]))
    print("Stuck buckets ............... %d" % len(snap["stuck_buckets"]))
    print("Detail rows ................. %d" % n_rows)
    print("Elapsed ..................... %.0f seconds" % (time.time() - t0))
    print("\nXLSX: %s" % xlsx_path)
    print("JSON: %s" % json_path)


if __name__ == "__main__":
    main()

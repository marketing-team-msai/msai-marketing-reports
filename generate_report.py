#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Marketing Pipeline Influence Report - deterministic generator.

Reproduces the 5-tab Excel workbook and adds an
ad-source (Windsor.ai) leg. Pure code: the entire data pull, workbook, chart,
runs with zero language-model calls in the hot path.

Universe: live HubSpot pull, portal 20335613. Deals created since 2025-06-01,
all pipelines, all stages. Influence = membership in the HubSpot "Campaign
Influence" folder segment lists. Attribution = even-split.

All credentials come from config.env (chmod 600, never web-served, never in git).
No em dashes anywhere.
"""
import os, sys, json, time, math
from collections import defaultdict
from datetime import datetime, timezone

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------- config ------
# Environment keys this module will accept as a stand-in for config.env.
# Scoped to known prefixes so os.environ at large never lands in CFG.
_ENV_PREFIXES = ("HUBSPOT_", "CAMPAIGN_", "DEALS_", "WINDSOR_", "NET_NEW_",
                 "SLA_DAYS_", "SUPABASE_", "MSAI_", "MKTG_")
_ENV_KEYS = ("PUBLISH_TARGET",)

def load_config(path=None):
    """Read config.env, then fill anything it does not set from the environment.

    The file is optional so this module can run on environment variables alone.
    When the file exists its values win, which keeps the CI path (the workflow
    writes config.env from secrets) behaving exactly as it did before."""
    path = path or os.path.join(HERE, "config.env")
    cfg = {}
    for k, v in os.environ.items():
        if k.startswith(_ENV_PREFIXES) or k in _ENV_KEYS:
            cfg[k] = v
    if os.path.exists(path):
        with open(path) as fh:
            for line in fh:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    cfg[k.strip()] = v.strip()
    return cfg

# Settings live here but are populated by init(), not at import time, so that
# importing this module for its pull/compute functions costs nothing and does
# not require a token to be present. main() calls init() first; so does any
# external caller (see sync_to_mktg.py).
CFG = {}
HS_TOKEN = None
PORTAL = "20335613"
FOLDER_ID = "240016304"
SINCE = "2025-06-01"
WINDSOR_KEY = ""
WINDSOR_PRESET = "last_365d"
OUTDIR = os.path.join(HERE, "output")
HS = "https://api.hubapi.com"
HH = {}
_READY = False

def init(path=None, make_outdir=True):
    """Load config and populate module settings. Idempotent."""
    global CFG, HS_TOKEN, PORTAL, FOLDER_ID, SINCE, WINDSOR_KEY, WINDSOR_PRESET
    global HH, _READY
    if _READY:
        return CFG
    CFG = load_config(path)
    HS_TOKEN = CFG.get("HUBSPOT_TOKEN") or ""
    if not HS_TOKEN:
        raise RuntimeError("HUBSPOT_TOKEN is not set. Put it in config.env next "
                           "to this script, or export it in the environment.")
    PORTAL = CFG.get("HUBSPOT_PORTAL_ID", "20335613")
    FOLDER_ID = CFG.get("CAMPAIGN_INFLUENCE_FOLDER_ID", "240016304")
    SINCE = CFG.get("DEALS_CREATED_SINCE", "2025-06-01")
    WINDSOR_KEY = CFG.get("WINDSOR_API_KEY", "")
    WINDSOR_PRESET = CFG.get("WINDSOR_DATE_PRESET", "last_365d")
    HH = {"Authorization": "Bearer " + HS_TOKEN, "Content-Type": "application/json"}
    if make_outdir:
        os.makedirs(OUTDIR, exist_ok=True)
    _READY = True
    return CFG

# ------------------------------------------------------------- hubspot io -----
def hs_get(path, params=None):
    for attempt in range(5):
        r = requests.get(HS + path, headers=HH, params=params, timeout=60)
        if r.status_code == 429:
            time.sleep(1.5 * (attempt + 1)); continue
        r.raise_for_status()
        return r.json()
    raise RuntimeError("rate limited: GET " + path)

def hs_post(path, body):
    for attempt in range(5):
        r = requests.post(HS + path, headers=HH, json=body, timeout=60)
        if r.status_code == 429:
            time.sleep(1.5 * (attempt + 1)); continue
        r.raise_for_status()
        return r.json()
    raise RuntimeError("rate limited: POST " + path)

def since_ms():
    dt = datetime.strptime(SINCE, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    return str(int(dt.timestamp() * 1000))

# ------------------------------------------------------- campaign classifier --
# Reverse-engineered from Alecia's July 9 workbook (name-keyword rules).
def classify_campaign(name):
    n = name.lower()
    if "pr:" in n:                                    return "PR"
    if "case study" in n or "case study:" in n:       return "Case Study"
    if "webinar" in n:                                return "Webinar"
    if "whitepaper" in n:                             return "Whitepaper"
    if "video" in n:                                  return "Video"
    if "blog" in n:                                   return "Blog"
    if ("conference" in n) or ("summit" in n) or ("live event" in n):
        return "Event"
    if ("web form" in n) or ("request a quote" in n) or ("solution brief" in n):
        return "Form"
    return "Content"

# ------------------------------------------------------------- data pull ------
def pull_pipelines():
    data = hs_get("/crm/v3/pipelines/deals")
    pipe_label, stage_label = {}, {}
    for p in data["results"]:
        pipe_label[p["id"]] = p["label"]
        for s in p["stages"]:
            stage_label[s["id"]] = s["label"]
    return pipe_label, stage_label

def pull_campaign_lists():
    """All lists whose name starts with 'Campaign Influence' (the folder %s)."""
    res = hs_post("/crm/v3/lists/search", {"query": "Campaign Influence", "count": 250})
    lists = [l for l in res.get("lists", [])
             if l["name"].lower().startswith("campaign influence")]
    return lists

def pull_memberships(list_id):
    ids, after = [], None
    while True:
        params = {"limit": 250}
        if after:
            params["after"] = after
        r = hs_get("/crm/v3/lists/%s/memberships" % list_id, params)
        ids.extend(str(m["recordId"]) for m in r.get("results", []))
        after = (r.get("paging", {}).get("next") or {}).get("after")
        if not after:
            break
    return ids

def pull_deals():
    ms = since_ms()
    props = ["dealname", "amount", "amount_in_home_currency", "pipeline", "dealstage",
             "closedate", "createdate", "hs_is_closed_won", "hs_is_closed"]
    out, after = [], None
    while True:
        body = {"filterGroups": [{"filters": [
                    {"propertyName": "createdate", "operator": "GTE", "value": ms}]}],
                "properties": props, "limit": 100, "sorts": [{"propertyName": "createdate", "direction": "ASCENDING"}]}
        if after:
            body["after"] = after
        r = hs_post("/crm/v3/objects/deals/search", body)
        out.extend(r.get("results", []))
        after = (r.get("paging", {}).get("next") or {}).get("after")
        if not after:
            break
        time.sleep(0.25)
    return out

def batch_assoc(from_obj, to_obj, ids):
    """v4 batch associations. Returns dict[fromId] -> [toId,...]."""
    out = defaultdict(list)
    ids = [str(i) for i in ids]
    for i in range(0, len(ids), 100):
        chunk = ids[i:i + 100]
        r = hs_post("/crm/v4/associations/%s/%s/batch/read" % (from_obj, to_obj),
                    {"inputs": [{"id": x} for x in chunk]})
        for res in r.get("results", []):
            fid = str((res.get("from") or {}).get("id") or res.get("fromObjectId"))
            for t in res.get("to", []):
                tid = t.get("toObjectId") or t.get("id")
                if tid is not None:
                    out[fid].append(str(tid))
        time.sleep(0.15)
    return out

def batch_read(obj, ids, properties):
    out = {}
    ids = [str(i) for i in ids]
    for i in range(0, len(ids), 100):
        chunk = ids[i:i + 100]
        r = hs_post("/crm/v3/objects/%s/batch/read" % obj,
                    {"properties": properties, "inputs": [{"id": x} for x in chunk]})
        for res in r.get("results", []):
            out[str(res["id"])] = res.get("properties", {})
        time.sleep(0.1)
    return out

# ----------------------------------------------------------------- compute ----
def build_dataset():
    print("[1/6] pipelines ...", flush=True)
    pipe_label, stage_label = pull_pipelines()

    print("[2/6] campaign influence lists ...", flush=True)
    lists = pull_campaign_lists()
    print("      %d campaign lists" % len(lists), flush=True)

    contact_campaigns = defaultdict(set)   # contactId -> {campaign name}
    campaign_size = {}                      # campaign name -> membership count
    for l in lists:
        name = l["name"]
        members = pull_memberships(l["listId"])
        campaign_size[name] = len(members)
        for c in members:
            contact_campaigns[c].add(name)
    total_contacts_mapped = len(contact_campaigns)
    print("      %d contacts mapped across lists" % total_contacts_mapped, flush=True)

    print("[3/6] deals ...", flush=True)
    raw_deals = pull_deals()
    deals = {}
    for d in raw_deals:
        p = d["properties"]
        # Home currency only, matching the rule used across this system. The
        # raw amount is the fallback when a deal has no home-currency value.
        amt = float(p.get("amount_in_home_currency") or p.get("amount") or 0)
        cd = (p.get("closedate") or "")[:10]
        deals[d["id"]] = {
            "id": d["id"], "name": p.get("dealname") or "",
            "amount": amt,
            "amount_home": amt,
            "pipeline": pipe_label.get(p.get("pipeline"), p.get("pipeline") or ""),
            "stage": stage_label.get(p.get("dealstage"), p.get("dealstage") or ""),
            "close": cd,
            "create": (p.get("createdate") or "")[:10],
            "won": str(p.get("hs_is_closed_won")).lower() == "true",
            "closed": str(p.get("hs_is_closed")).lower() == "true",
        }
    total_deals = len(deals)
    total_value = sum(x["amount"] for x in deals.values())
    print("      %d deals, $%.2f total value" % (total_deals, total_value), flush=True)

    print("[4/6] deal associations (contacts, companies) ...", flush=True)
    deal_ids = list(deals.keys())
    d2c = batch_assoc("deals", "contacts", deal_ids)
    d2co = batch_assoc("deals", "companies", deal_ids)

    # influenced deals + (deal x influenced contact) rows
    influenced_deal_ids = set()
    pair_rows = []  # (deal, contactId)
    contacts_needed = set()
    for did, d in deals.items():
        inf_contacts = [c for c in d2c.get(did, []) if c in contact_campaigns]
        if inf_contacts:
            influenced_deal_ids.add(did)
            for c in sorted(set(inf_contacts)):
                pair_rows.append((did, c))
                contacts_needed.add(c)

    # multi-touch contacts: associated with a deal in the window AND in 2+ lists.
    # (Scoped to deal-associated contacts. The full mapped universe overlaps
    # broadly because some lists are large; the diagnostic tracks contacts that
    # actually reach a deal touched by more than one campaign.)
    deal_contact_ids = set()
    for clist in d2c.values():
        deal_contact_ids.update(clist)
    multi_ids = [c for c in deal_contact_ids
                 if c in contact_campaigns and len(contact_campaigns[c]) >= 2]
    contacts_needed.update(multi_ids)

    print("[5/6] contact + company details ...", flush=True)
    companies_needed = set()
    for did in influenced_deal_ids:
        companies_needed.update(d2co.get(did, []))
    # also companies for multi-touch contacts' deals
    contact_deals = defaultdict(list)  # contactId -> [dealId] (window deals associated)
    for did, clist in d2c.items():
        for c in clist:
            contact_deals[c].append(did)

    contact_props = batch_read("contacts", contacts_needed,
                               ["firstname", "lastname", "email"]) if contacts_needed else {}
    company_props = batch_read("companies", companies_needed, ["name"]) if companies_needed else {}

    def cname(cid):
        p = contact_props.get(cid, {})
        nm = ((p.get("firstname") or "") + " " + (p.get("lastname") or "")).strip()
        return nm or (p.get("email") or cid)
    def cemail(cid):
        return contact_props.get(cid, {}).get("email") or ""
    def deal_company_name(did):
        cos = d2co.get(did, [])
        if cos:
            return company_props.get(cos[0], {}).get("name") or ""
        return ""
    def camps_str(cid):
        return " | ".join(sorted(contact_campaigns[cid], key=str.lower))

    influenced_value = sum(deals[d]["amount"] for d in influenced_deal_ids)

    return dict(
        pipe_label=pipe_label, stage_label=stage_label, lists=lists,
        contact_campaigns=contact_campaigns, campaign_size=campaign_size,
        total_contacts_mapped=total_contacts_mapped,
        deals=deals, total_deals=total_deals, total_value=total_value,
        d2c=d2c, d2co=d2co,
        influenced_deal_ids=influenced_deal_ids, influenced_value=influenced_value,
        pair_rows=pair_rows, multi_ids=multi_ids, contact_deals=contact_deals,
        contact_props=contact_props, company_props=company_props,
        cname=cname, cemail=cemail, deal_company_name=deal_company_name, camps_str=camps_str,
    )

def compute_campaign_summary(ds):
    """Per-deal-union even-split (Alecia's proven method).

    For each influenced deal, take the UNION of campaigns across ALL of that
    deal's influenced contacts, then add deal_amount / len(union) ONCE to each
    campaign in the union. This de-double-counts deals that have 2+ influenced
    contacts, so column E sums exactly to the de-duplicated influenced total
    (a deal contributes its full amount once, spread across its campaign union).

    Columns C (Influenced Contacts) and D (Pipeline Deals Touched) are unchanged:
    C is raw list membership; D is the count of influenced deals a campaign
    touches, which the union loop reproduces identically."""
    camp_value = defaultdict(float)
    camp_deals = defaultdict(set)
    deals = ds["deals"]; cc = ds["contact_campaigns"]
    # group each influenced deal's influenced contacts together
    deal_contacts = defaultdict(set)
    for did, cid in ds["pair_rows"]:
        deal_contacts[did].add(cid)
    for did, cids in deal_contacts.items():
        union = set()
        for cid in cids:
            union |= cc[cid]
        if not union:
            continue
        share = deals[did]["amount"] / len(union)
        for camp in union:
            camp_value[camp] += share
            camp_deals[camp].add(did)
    rows = []
    for name, size in ds["campaign_size"].items():
        deals_touched = len(camp_deals.get(name, ()))
        rows.append({
            "name": name, "type": classify_campaign(name),
            "contacts": size,
            "deals": deals_touched,
            "value": camp_value.get(name, 0.0),
        })
    # producing (deals>0) sorted by value desc; non-producing by contacts desc
    producing = sorted([r for r in rows if r["deals"] > 0], key=lambda r: -r["value"])
    idle = sorted([r for r in rows if r["deals"] == 0], key=lambda r: -r["contacts"])
    return producing + idle, len(producing)

def compute_multitouch(ds):
    rows = []
    for cid in ds["multi_ids"]:
        camps = sorted(ds["contact_campaigns"][cid], key=str.lower)
        window_deals = [d for d in set(ds["contact_deals"].get(cid, []))
                        if d in ds["influenced_deal_ids"]]
        pipeline = sum(ds["deals"][d]["amount"] for d in window_deals)
        companies = sorted({ds["deal_company_name"](d) for d in window_deals if ds["deal_company_name"](d)})
        rows.append({
            "name": ds["cname"](cid), "email": ds["cemail"](cid),
            "ncamps": len(camps), "camps": " | ".join(camps),
            "ndeals": len(window_deals), "pipeline": pipeline,
            "companies": " | ".join(companies),
        })
    rows.sort(key=lambda r: (-r["ncamps"], -r["pipeline"], r["name"].lower()))
    return rows

# ----------------------------------------------------------------- windsor ----
def pull_windsor():
    if not WINDSOR_KEY:
        return None
    fields = "source,account_name,campaign,clicks,spend,impressions,conversions"
    url = ("https://connectors.windsor.ai/all?api_key=%s&fields=%s&date_preset=%s"
           % (WINDSOR_KEY, fields, WINDSOR_PRESET))
    try:
        r = requests.get(url, timeout=90)
        r.raise_for_status()
        data = r.json()
        rows = data.get("data", data if isinstance(data, list) else [])
    except Exception as e:
        print("      windsor pull failed: %s" % str(e)[:120], flush=True)
        return None
    agg = defaultdict(lambda: {"account": "", "clicks": 0.0, "spend": 0.0,
                               "impr": 0, "conv": 0.0})
    for row in rows:
        s = row.get("source", "?")
        a = agg[s]
        a["account"] = row.get("account_name") or a["account"]
        a["clicks"] += float(row.get("clicks") or 0)
        a["spend"] += float(row.get("spend") or 0)
        a["impr"] += int(float(row.get("impressions") or 0))
        a["conv"] += float(row.get("conversions") or 0)
    return agg

# =============================================================== EXCEL =========
C_TITLE = PatternFill("solid", fgColor="1F4E78")
C_HDR = PatternFill("solid", fgColor="2E75B6")
F_TITLE = Font(bold=True, size=16, color="FFFFFF")
F_HDR = Font(bold=True, size=11, color="FFFFFF")
F_KPI_HDR = Font(bold=True, size=10, color="595959")
F_KPI_VAL = Font(bold=True, size=18, color="1F4E78")
F_SECTION = Font(bold=True, size=14, color="1F4E78")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
USD = '"$"#,##0'

def _hdr_row(ws, row, headers, widths=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.fill = C_HDR; cell.font = F_HDR
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    if widths:
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

def build_workbook(ds, camp_rows, n_producing, multi_rows, windsor, gen_date):
    wb = openpyxl.Workbook()

    # ---- Tab 1: Executive Summary ----
    ws = wb.active; ws.title = "Executive Summary"
    ws.cell(1, 1, "MSAI Marketing Pipeline Influence Report").font = F_TITLE
    ws.cell(1, 1).fill = C_TITLE
    for c in range(2, 7):
        ws.cell(1, c).fill = C_TITLE
    meta = ("Generated: %s  |  Deals: live HubSpot pull  |  Created since %s  |  "
            "All pipelines · All stages  |  Influence source: 'Campaign Influence' folder lists"
            % (gen_date.strftime("%B %d, %Y"), SINCE))
    ws.cell(2, 1, meta).font = Font(italic=True, size=9, color="595959")
    kpi_hdr = ["Total Deals (all pipelines)", "Total Pipeline Value",
               "Deals w/ Mktg Influence", "Influenced Pipeline Value",
               "Campaigns Checked", "Total Contacts Mapped"]
    for c, h in enumerate(kpi_hdr, 1):
        cell = ws.cell(4, c, h); cell.font = F_KPI_HDR
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    inf_pct = (ds["influenced_value"] / ds["total_value"] * 100) if ds["total_value"] else 0
    vals = [ds["total_deals"], ds["total_value"], len(ds["influenced_deal_ids"]),
            ds["influenced_value"], len(ds["campaign_size"]), ds["total_contacts_mapped"]]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(5, c, v); cell.font = F_KPI_VAL
        if c in (2, 4):
            cell.number_format = USD
        elif c in (1, 3, 5, 6):
            cell.number_format = '#,##0'
    ws.cell(6, 4, "%.1f%% of total pipeline value" % inf_pct).font = Font(italic=True, size=9, color="595959")
    ws.cell(8, 1, "HOW TO READ THIS REPORT").font = F_SECTION
    how = [
        ("Tab 2 – Campaign Influence Map",
         "Every (deal x influenced contact) pair. Shows which campaign(s) touched which contact on each open or closed deal in the window."),
        ("Tab 3 – Campaigns Summary",
         "All campaigns ranked by pipeline value influenced. Even-split attribution: deal value divided across campaigns that touched any associated contact."),
        ("Tab 4 – Deal Contact Details",
         "Full filterable contact-level view of every campaign-influenced contact with their associated deals and campaign attribution."),
        ("Tab 5 – Multi-Touch Contacts",
         "Diagnostic. Lists contacts touched by 2+ campaigns. As nurture flows and retargeting fire, this should grow."),
        ("Attribution method",
         "A deal counts as 'influenced' if any associated contact is a member of any list in the Campaign Influence folder. Pipeline value is split evenly across all influencing campaigns to avoid double-counting in tab 3 totals."),
        ("Data note",
         "HubSpot Pro accounts cannot use native campaign attribution APIs. This report uses manually-maintained 'Campaign Influence' segment lists in HubSpot (folder ID %s) as a proxy for campaign engagement." % FOLDER_ID),
    ]
    for i, (a, b) in enumerate(how):
        r = 9 + i
        ws.cell(r, 1, a).font = Font(bold=True, size=10)
        ws.cell(r, 2, b).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 28
    for col in "BCDEF":
        ws.column_dimensions[col].width = 22

    # ---- Tab 2: Campaign Influence Map ----
    ws = wb.create_sheet("Campaign Influence Map")
    heads = ["Deal Name", "Company Name", "Pipeline", "Stage", "Amount ($)", "Close Date",
             "Contact Name", "Email", "Campaign(s) Influencing Contact", "Influenced?"]
    _hdr_row(ws, 1, heads, [42, 28, 22, 22, 14, 12, 22, 30, 50, 14])
    # sort deals by amount desc, then rows by contact name
    ordered = sorted(ds["pair_rows"],
                     key=lambda pc: (-ds["deals"][pc[0]]["amount"], ds["cname"](pc[1]).lower()))
    r = 2
    for did, cid in ordered:
        d = ds["deals"][did]
        ws.cell(r, 1, d["name"]); ws.cell(r, 2, ds["deal_company_name"](did))
        ws.cell(r, 3, d["pipeline"]); ws.cell(r, 4, d["stage"])
        ws.cell(r, 5, d["amount"]).number_format = USD
        ws.cell(r, 6, d["close"]); ws.cell(r, 7, ds["cname"](cid))
        ws.cell(r, 8, ds["cemail"](cid)); ws.cell(r, 9, ds["camps_str"](cid))
        ws.cell(r, 10, "✓ Influenced")
        r += 1
    ws.freeze_panes = "A2"; ws.auto_filter.ref = "A1:J%d" % (r - 1)

    # ---- Tab 3: Campaigns Summary ----
    ws = wb.create_sheet("Campaigns Summary")
    heads = ["Campaign Name", "Type", "Influenced Contacts (total)",
             "Pipeline Deals Touched", "Pipeline Value Influenced (even-split)"]
    _hdr_row(ws, 1, heads, [70, 14, 22, 22, 30])
    r = 2
    for row in camp_rows:
        ws.cell(r, 1, row["name"]); ws.cell(r, 2, row["type"])
        ws.cell(r, 3, row["contacts"])
        if row["deals"] > 0:
            ws.cell(r, 4, row["deals"])
            ws.cell(r, 5, row["value"]).number_format = USD
        else:
            ws.cell(r, 4, "—"); ws.cell(r, 5, "—")
        r += 1
    ws.freeze_panes = "A2"; ws.auto_filter.ref = "A1:E%d" % (r - 1)
    r += 1
    ws.cell(r, 1, "TOTAL (de-duplicated)").font = Font(bold=True)
    ws.cell(r, 4, len(ds["influenced_deal_ids"])).font = Font(bold=True)
    tc = ws.cell(r, 5, ds["influenced_value"]); tc.font = Font(bold=True); tc.number_format = USD
    ws.cell(r + 1, 1, "Note: column E uses per-deal-union even-split. Each influenced deal's "
            "value is divided once across the union of campaigns that touched it, so column E "
            "sums to the de-duplicated influenced total in the TOTAL row.").font = \
        Font(italic=True, size=9, color="595959")

    # ---- Tab 4: Deal Contact Details ----
    ws = wb.create_sheet("Deal Contact Details")
    heads = ["Deal Name", "Company Name", "Contact Name", "Email", "Stage",
             "Deal Amount ($)", "Campaign(s) Influencing Contact", "Mktg Influenced?"]
    _hdr_row(ws, 1, heads, [42, 28, 22, 30, 22, 16, 50, 14])
    r = 2
    for did, cid in ordered:
        d = ds["deals"][did]
        ws.cell(r, 1, d["name"]); ws.cell(r, 2, ds["deal_company_name"](did))
        ws.cell(r, 3, ds["cname"](cid)); ws.cell(r, 4, ds["cemail"](cid))
        ws.cell(r, 5, d["stage"]); ws.cell(r, 6, d["amount"]).number_format = USD
        ws.cell(r, 7, ds["camps_str"](cid)); ws.cell(r, 8, "Yes")
        r += 1
    ws.freeze_panes = "A2"; ws.auto_filter.ref = "A1:H%d" % (r - 1)

    # ---- Tab 5: Multi-Touch Contacts ----
    ws = wb.create_sheet("Multi-Touch Contacts")
    ws.cell(1, 1, "Multi-Touch Contacts — Diagnostic").font = Font(bold=True, size=14, color="1F4E78")
    ws.cell(2, 1, "Contacts in 2+ Campaign Influence lists. As nurture sequences and retargeting "
            "fire, this list should grow, meaning the same contact is touched by multiple campaigns "
            "before becoming a deal. A healthy program produces a long, multi-line tail.").alignment = \
        Alignment(wrap_text=True)
    heads = ["Contact Name", "Email", "# Campaigns", "Campaigns", "# Deals",
             "Influenced Pipeline ($)", "Companies"]
    _hdr_row(ws, 4, heads, [22, 30, 14, 60, 12, 22, 35])
    r = 5
    for m in multi_rows:
        ws.cell(r, 1, m["name"]); ws.cell(r, 2, m["email"]); ws.cell(r, 3, m["ncamps"])
        ws.cell(r, 4, m["camps"]); ws.cell(r, 5, m["ndeals"])
        ws.cell(r, 6, m["pipeline"]).number_format = USD; ws.cell(r, 7, m["companies"])
        r += 1
    ws.freeze_panes = "A5"
    r += 1
    ws.cell(r, 1, "TOTAL multi-touch contacts:").font = Font(bold=True)
    ws.cell(r, 2, len(multi_rows)).font = Font(bold=True)
    maxt = max((m["ncamps"] for m in multi_rows), default=0)
    ws.cell(r + 1, 1, "Max campaign touches on any contact:").font = Font(bold=True)
    ws.cell(r + 1, 2, maxt).font = Font(bold=True)

    # ---- Tab 6: Ad Source Performance (Windsor.ai) ----
    ws = wb.create_sheet("Ad Source Performance")
    ws.cell(1, 1, "Ad Source Performance — Windsor.ai").font = Font(bold=True, size=14, color="1F4E78")
    ws.cell(2, 1, "Ad-source pull (window: %s). Facebook is not connected. Split into PAID MEDIA "
            "(sources with ad spend) and ORGANIC / REFERRAL (zero-spend GA4 and Search Console "
            "sources such as bing, chatgpt, (direct), duckduckgo, (not set)). The paid total is "
            "the ad-performance number: zero-spend organic conversions are kept separate so the "
            "paid Reported Conversions total is not inflated." % WINDSOR_PRESET).alignment = \
        Alignment(wrap_text=True)
    heads = ["Source", "Account", "Clicks", "Impressions", "Spend ($)", "Reported Conversions"]
    _hdr_row(ws, 4, heads, [20, 34, 14, 16, 16, 20])
    r = 5
    if windsor:
        paid_rows = {s: a for s, a in windsor.items() if a["spend"] > 0}
        organic_rows = {s: a for s, a in windsor.items() if a["spend"] <= 0}

        def _emit(rows_dict):
            nonlocal r
            sp = cl = cv = 0.0
            keyf = lambda x: (-x[1]["spend"], -x[1]["clicks"])
            for s, a in sorted(rows_dict.items(), key=keyf):
                ws.cell(r, 1, s); ws.cell(r, 2, a["account"]); ws.cell(r, 3, int(a["clicks"]))
                ws.cell(r, 4, a["impr"]); ws.cell(r, 5, a["spend"]).number_format = USD
                ws.cell(r, 6, int(a["conv"])); r += 1
                sp += a["spend"]; cl += a["clicks"]; cv += a["conv"]
            return sp, cl, cv

        def _subtotal(label, sp, cl, cv, color="1F4E78"):
            nonlocal r
            ws.cell(r, 1, label).font = Font(bold=True, color=color)
            ws.cell(r, 3, int(cl)).font = Font(bold=True)
            c5 = ws.cell(r, 5, sp); c5.font = Font(bold=True); c5.number_format = USD
            ws.cell(r, 6, int(cv)).font = Font(bold=True)
            r += 1

        ws.cell(r, 1, "PAID MEDIA").font = Font(bold=True, size=11, color="1F4E78"); r += 1
        p_sp, p_cl, p_cv = _emit(paid_rows)
        _subtotal("Paid total", p_sp, p_cl, p_cv)
        r += 1
        ws.cell(r, 1, "ORGANIC / REFERRAL (no ad spend)").font = Font(bold=True, size=11, color="595959"); r += 1
        ws.cell(r, 1, "Zero-spend sources; excluded from the paid conversions total, shown for completeness.").font = \
            Font(italic=True, size=9, color="595959"); r += 1
        o_sp, o_cl, o_cv = _emit(organic_rows)
        _subtotal("Organic total", o_sp, o_cl, o_cv, color="595959")
        r += 1
        _subtotal("GRAND TOTAL (paid + organic)", p_sp + o_sp, p_cl + o_cl, p_cv + o_cv)
        ws.cell(r + 1, 1, "Watch: paid spend of $%s across Google, LinkedIn and Reddit produced only "
                "%d reported ad conversions in this window (paid total above). Reconcile ad platforms "
                "against HubSpot influenced pipeline before the next flight."
                % ("{:,.0f}".format(p_sp), int(p_cv))).font = \
            Font(italic=True, size=10, color="C00000")
    else:
        ws.cell(r, 1, "Windsor.ai pull unavailable this run.")
    ws.freeze_panes = "A5"

    path = os.path.join(OUTDIR, "MSAI_Pipeline_Influence_Report.xlsx")
    wb.save(path)
    return path

# =============================================================== CHARTS ========
# =============================================================== MAIN ==========
def main():
    init()
    gen_date = datetime.now(timezone.utc)
    # prior-refresh snapshot for week-over-week (seeded with the July 9 baseline)
    prior_path = os.path.join(HERE, "prior_snapshot.json")
    default_prior = {"label": "July 09, 2026", "total_deals": 931, "total_value": 42563459.06,
                     "influenced_deals": 193, "influenced_value": 9988074.95,
                     "producing": 24, "multitouch": 46, "contacts_mapped": 5505}
    prior = json.load(open(prior_path)) if os.path.exists(prior_path) else default_prior

    ds = build_dataset()
    print("[6/6] windsor + assembly ...", flush=True)
    windsor = pull_windsor()
    camp_rows, n_producing = compute_campaign_summary(ds)
    multi_rows = compute_multitouch(ds)

    xlsx = build_workbook(ds, camp_rows, n_producing, multi_rows, windsor, gen_date)

    # write current snapshot for next week's WoW
    snap = {"label": gen_date.strftime("%B %d, %Y"), "total_deals": ds["total_deals"],
            "total_value": ds["total_value"], "influenced_deals": len(ds["influenced_deal_ids"]),
            "influenced_value": ds["influenced_value"], "producing": n_producing,
            "multitouch": len(multi_rows), "contacts_mapped": ds["total_contacts_mapped"]}
    json.dump(snap, open(prior_path, "w"), indent=2)

    # machine-readable metrics for the Supabase history push (push_history.py).
    # Same numbers as the workbook exec summary, nothing recomputed.
    metrics = {
        "run_date": gen_date.strftime("%Y-%m-%d"),
        "generated_at": gen_date.isoformat(),
        "headline": snap,
        "campaigns": [
            {"campaign_name": r["name"], "campaign_type": r["type"],
             "contacts": r["contacts"], "deals": r["deals"], "value": r["value"]}
            for r in camp_rows
        ],
        "windsor_paid_spend": (
            sum(a["spend"] for k, a in windsor.items() if k != "searchconsole")
            if windsor else None
        ),
    }
    with open(os.path.join(OUTDIR, "influence_metrics.json"), "w") as fh:
        json.dump(metrics, fh, indent=2)

    inf_pct = (ds["influenced_value"] / ds["total_value"] * 100) if ds["total_value"] else 0
    print("\n================ HEADLINE KPIs ================")
    print("Total Deals ................. {:,}".format(ds["total_deals"]))
    print("Total Pipeline Value ........ ${:,.2f}".format(ds["total_value"]))
    print("Deals w/ Mktg Influence ..... %d (%.1f%% of value)" % (len(ds["influenced_deal_ids"]), inf_pct))
    print("Influenced Pipeline Value ... ${:,.2f}".format(ds["influenced_value"]))
    print("Campaigns Checked ........... %d (producing: %d)" % (len(ds["campaign_size"]), n_producing))
    print("Total Contacts Mapped ....... {:,}".format(ds["total_contacts_mapped"]))
    print("Multi-touch contacts ........ %d" % len(multi_rows))
    print("Tab2/Tab4 (deal x contact) .. %d rows" % len(ds["pair_rows"]))
    if windsor:
        paid = sum(a["spend"] for s, a in windsor.items() if s != "searchconsole")
        print("Windsor paid spend ({}) ... ${:,.2f}".format(WINDSOR_PRESET, paid))
    print("\nXLSX: %s" % xlsx)

if __name__ == "__main__":
    main()
